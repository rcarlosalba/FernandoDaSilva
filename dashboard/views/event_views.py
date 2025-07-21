import csv
import io
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.utils import timezone
from constants.constant import manager_required
from events.models import Event, Category, PaymentMethod, Registration, Payment, Survey, SurveyResponse
from events.forms import EventForm, CategoryForm, PaymentMethodForm, SurveyForm, SurveyQuestionFormSet, SurveyQuestionOptionFormSet

# ===== VISTAS DE EVENTOS (GESTIÓN) =====


@manager_required
def event_list(request):
    """
    Lista de eventos para el dashboard.
    """
    events = Event.objects.select_related(
        'created_by').prefetch_related('categories').all()

    # Filtros
    status = request.GET.get('status')
    if status:
        events = events.filter(status=status)

    event_type = request.GET.get('event_type')
    if event_type:
        events = events.filter(event_type=event_type)

    search = request.GET.get('search')
    if search:
        events = events.filter(
            Q(title__icontains=search) |
            Q(description__icontains=search) |
            Q(location__icontains=search)
        )

    # Estadísticas
    total_events = events.count()
    published_events = events.filter(status='published').count()
    draft_events = events.filter(status='draft').count()
    upcoming_events = events.filter(start_date__gt=timezone.now()).count()

    # Paginación
    paginator = Paginator(events, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'events': page_obj,
        'total_events': total_events,
        'published_events': published_events,
        'draft_events': draft_events,
        'upcoming_events': upcoming_events,
    }

    return render(request, 'dashboard/events/event_list.html', context)


@manager_required
def event_create(request):
    """
    Crear nuevo evento.
    """
    if request.method == 'POST':
        form = EventForm(request.POST, request.FILES)
        if form.is_valid():
            event = form.save(commit=False)
            event.created_by = request.user
            event.save()
            form.save_m2m()  # Guardar relaciones ManyToMany

            messages.success(
                request, f'Evento "{event.title}" creado exitosamente.')
            return redirect('dashboard:event_detail', pk=event.pk)
    else:
        form = EventForm()

    context = {
        'form': form,
        'title': 'Nuevo Evento',
        'action': 'Crear Evento'
    }

    return render(request, 'dashboard/events/event_form.html', context)


@manager_required
def event_detail(request, pk):
    """
    Detalle de evento.
    """
    event = get_object_or_404(Event.objects.select_related(
        'created_by').prefetch_related('categories', 'payment_methods'), pk=pk)

    # Obtener inscripciones con pagos
    registrations = event.registrations.select_related('payment').all()

    # Estadísticas
    total_registrations = registrations.count()
    accepted_registrations = registrations.filter(status='accepted').count()
    pending_registrations = registrations.filter(status='pending').count()
    waitlist_registrations = registrations.filter(status='waitlist').count()

    context = {
        'event': event,
        'registrations': registrations,
        'total_registrations': total_registrations,
        'accepted_registrations': accepted_registrations,
        'pending_registrations': pending_registrations,
        'waitlist_registrations': waitlist_registrations,
    }

    return render(request, 'dashboard/events/event_detail.html', context)


@manager_required
def event_update(request, pk):
    """
    Editar evento.
    """
    event = get_object_or_404(Event, pk=pk)

    if request.method == 'POST':
        form = EventForm(request.POST, request.FILES, instance=event)
        if form.is_valid():
            form.save()
            messages.success(
                request, f'Evento "{event.title}" actualizado exitosamente.')
            return redirect('dashboard:event_detail', pk=event.pk)
    else:
        form = EventForm(instance=event)

    context = {
        'form': form,
        'object': event,
        'title': 'Editar Evento',
        'action': 'Guardar Cambios'
    }

    return render(request, 'dashboard/events/event_form.html', context)


@manager_required
def event_delete(request, pk):
    """
    Eliminar evento.
    """
    event = get_object_or_404(Event, pk=pk)

    if request.method == 'POST':
        event_title = event.title
        event.delete()
        messages.success(
            request, f'Evento "{event_title}" eliminado exitosamente.')
        return redirect('dashboard:event_list')

    context = {
        'event': event
    }

    return render(request, 'dashboard/events/event_delete.html', context)


# ===== VISTAS DE CATEGORÍAS =====

@manager_required
def category_list(request):
    """
    Lista de categorías.
    """
    categories = Category.objects.annotate(
        annotated_event_count=Count('events')).all()

    # Paginación
    paginator = Paginator(categories, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'categories': page_obj,
    }

    return render(request, 'dashboard/events/category_list.html', context)


@manager_required
def category_create(request):
    """
    Crear nueva categoría.
    """
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            category = form.save()
            messages.success(
                request, f'Categoría "{category.name}" creada exitosamente.')
            return redirect('dashboard:event_category_detail', pk=category.pk)
    else:
        form = CategoryForm()

    context = {
        'form': form,
        'title': 'Nueva Categoría',
        'action': 'Crear Categoría'
    }

    return render(request, 'dashboard/events/category_form.html', context)


@manager_required
def category_detail(request, pk):
    """
    Detalle de categoría.
    """
    category = get_object_or_404(Category.objects.annotate(
        annotated_event_count=Count('events')), pk=pk)
    events = category.events.all()[:10]  # Últimos 10 eventos

    context = {
        'category': category,
        'events': events,
    }

    return render(request, 'dashboard/events/category_detail.html', context)


@manager_required
def category_update(request, pk):
    """
    Editar categoría.
    """
    category = get_object_or_404(Category, pk=pk)

    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(
                request, f'Categoría "{category.name}" actualizada exitosamente.')
            return redirect('dashboard:event_category_detail', pk=category.pk)
    else:
        form = CategoryForm(instance=category)

    context = {
        'form': form,
        'category': category,
        'title': 'Editar Categoría',
        'action': 'Guardar Cambios'
    }

    return render(request, 'dashboard/events/category_form.html', context)


@manager_required
def category_delete(request, pk):
    """
    Eliminar categoría.
    """
    category = get_object_or_404(Category, pk=pk)

    if request.method == 'POST':
        category_name = category.name
        category.delete()
        messages.success(
            request, f'Categoría "{category_name}" eliminada exitosamente.')
        return redirect('dashboard:event_category_list')

    context = {
        'category': category
    }

    return render(request, 'dashboard/events/category_delete.html', context)


# ===== VISTAS DE MÉTODOS DE PAGO =====

@manager_required
def payment_method_list(request):
    """
    Lista de métodos de pago.
    """
    payment_methods = PaymentMethod.objects.all()

    # Paginación
    paginator = Paginator(payment_methods, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'payment_methods': page_obj,
    }

    return render(request, 'dashboard/events/payment_method_list.html', context)


@manager_required
def payment_method_create(request):
    """
    Crear nuevo método de pago.
    """
    if request.method == 'POST':
        form = PaymentMethodForm(request.POST)
        if form.is_valid():
            payment_method = form.save()
            messages.success(
                request, f'Método de pago "{payment_method.name}" creado exitosamente.')
            return redirect('dashboard:payment_method_detail', pk=payment_method.pk)
    else:
        form = PaymentMethodForm()

    context = {
        'form': form,
        'title': 'Nuevo Método de Pago',
        'action': 'Crear Método de Pago'
    }

    return render(request, 'dashboard/events/payment_method_form.html', context)


@manager_required
def payment_method_detail(request, pk):
    """
    Detalle de método de pago.
    """
    payment_method = get_object_or_404(PaymentMethod, pk=pk)
    events = payment_method.events.all()

    context = {
        'payment_method': payment_method,
        'events': events,
    }

    return render(request, 'dashboard/events/payment_method_detail.html', context)


@manager_required
def payment_method_update(request, pk):
    """
    Editar método de pago.
    """
    payment_method = get_object_or_404(PaymentMethod, pk=pk)

    if request.method == 'POST':
        form = PaymentMethodForm(request.POST, instance=payment_method)
        if form.is_valid():
            form.save()
            messages.success(
                request, f'Método de pago "{payment_method.name}" actualizado exitosamente.')
            return redirect('dashboard:payment_method_detail', pk=payment_method.pk)
    else:
        form = PaymentMethodForm(instance=payment_method)

    context = {
        'form': form,
        'payment_method': payment_method,
        'title': 'Editar Método de Pago',
        'action': 'Guardar Cambios'
    }

    return render(request, 'dashboard/events/payment_method_form.html', context)


@manager_required
def payment_method_delete(request, pk):
    """
    Eliminar método de pago.
    """
    payment_method = get_object_or_404(PaymentMethod, pk=pk)

    if request.method == 'POST':
        payment_method_name = payment_method.name
        payment_method.delete()
        messages.success(
            request, f'Método de pago "{payment_method_name}" eliminado exitosamente.')
        return redirect('dashboard:payment_method_list')

    context = {
        'payment_method': payment_method
    }

    return render(request, 'dashboard/events/payment_method_delete.html', context)


# ===== VISTAS DE INSCRIPCIONES =====

@manager_required
def registration_list(request):
    """
    Lista de inscripciones.
    """
    registrations = Registration.objects.select_related(
        'event', 'payment').all()

    # Filtros
    status = request.GET.get('status')
    if status:
        registrations = registrations.filter(status=status)

    event_id = request.GET.get('event')
    if event_id:
        registrations = registrations.filter(event_id=event_id)

    search = request.GET.get('search')
    if search:
        registrations = registrations.filter(
            Q(full_name__icontains=search) |
            Q(email__icontains=search) |
            Q(event__title__icontains=search)
        )

    # Estadísticas
    total_registrations = registrations.count()
    pending_registrations = registrations.filter(status='pending').count()
    accepted_registrations = registrations.filter(status='accepted').count()
    waitlist_registrations = registrations.filter(status='waitlist').count()

    # Paginación
    paginator = Paginator(registrations, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'registrations': page_obj,
        'total_registrations': total_registrations,
        'pending_registrations': pending_registrations,
        'accepted_registrations': accepted_registrations,
        'waitlist_registrations': waitlist_registrations,
        'events': Event.objects.all(),  # Para filtros
    }

    return render(request, 'dashboard/events/registration_list.html', context)


@manager_required
def registration_detail(request, pk):
    """
    Detalle de inscripción.
    """
    registration = get_object_or_404(
        Registration.objects.select_related('event', 'payment'), pk=pk)

    context = {
        'registration': registration,
    }

    return render(request, 'dashboard/events/registration_detail.html', context)


@manager_required
def registration_approve(request, pk):
    """
    Aprobar inscripción.
    """
    registration = get_object_or_404(Registration, pk=pk)

    if registration.status == 'pending':
        registration.status = 'accepted'
        registration.save()

        # Enviar email de aprobación
        try:
            send_registration_approved_email(registration)
        except Exception as e:
            # Log error pero no fallar la aprobación
            print(f"Error sending approval email: {e}")

        messages.success(
            request, f'Inscripción de {registration.full_name} aprobada.')
    else:
        messages.error(
            request, 'Solo se pueden aprobar inscripciones pendientes.')

    return redirect('dashboard:registration_detail', pk=registration.pk)


@manager_required
def registration_reject(request, pk):
    """
    Rechazar inscripción.
    """
    registration = get_object_or_404(Registration, pk=pk)

    if registration.status == 'pending':
        registration.status = 'rejected'
        registration.save()

        # Enviar email de rechazo
        try:
            send_registration_rejected_email(registration)
        except Exception as e:
            # Log error pero no fallar el rechazo
            print(f"Error sending rejection email: {e}")

        messages.success(
            request, f'Inscripción de {registration.full_name} rechazada.')
    else:
        messages.error(
            request, 'Solo se pueden rechazar inscripciones pendientes.')

    return redirect('dashboard:registration_detail', pk=registration.pk)


@manager_required
def payment_verify(request, pk):
    """
    Verificar pago.
    """
    payment = get_object_or_404(Payment, pk=pk)

    if payment.status == 'pending':
        # Verificar si la inscripción estaba pendiente antes
        was_pending = payment.registration.status == 'pending'

        payment.verify_payment(request.user)

        # Si la inscripción estaba pendiente y ahora está aceptada, enviar email
        if was_pending and payment.registration.status == 'accepted':
            try:
                send_registration_approved_email(payment.registration)
            except Exception as e:
                # Log error pero no fallar la verificación
                print(
                    f"Error sending approval email after payment verification: {e}")

        messages.success(
            request, f'Pago de {payment.registration.full_name} verificado.')
    else:
        messages.error(request, 'Solo se pueden verificar pagos pendientes.')

    return redirect('dashboard:registration_detail', pk=payment.registration.pk)


# ===== VISTAS DE ESTADÍSTICAS =====

@manager_required
def event_statistics(request):
    """
    Estadísticas generales de eventos.
    """
    # Estadísticas de eventos
    total_events = Event.objects.count()
    published_events = Event.objects.filter(status='published').count()
    upcoming_events = Event.objects.filter(
        start_date__gt=timezone.now()).count()
    finished_events = Event.objects.filter(end_date__lt=timezone.now()).count()

    # Estadísticas de inscripciones
    total_registrations = Registration.objects.count()
    pending_registrations = Registration.objects.filter(
        status='pending').count()
    accepted_registrations = Registration.objects.filter(
        status='accepted').count()

    # Estadísticas de pagos
    total_payments = Payment.objects.count()
    pending_payments = Payment.objects.filter(status='pending').count()
    verified_payments = Payment.objects.filter(status='verified').count()

    # Estadísticas de encuestas
    total_surveys = Survey.objects.count()
    active_surveys = Survey.objects.filter(status='active').count()
    total_survey_responses = SurveyResponse.objects.count()
    completed_survey_responses = SurveyResponse.objects.filter(
        status='completed').count()
    events_with_surveys = Event.objects.filter(survey__isnull=False).count()

    # Eventos más populares
    popular_events = Event.objects.annotate(
        registration_count=Count('registrations')
    ).order_by('-registration_count')[:5]

    # Categorías más populares
    popular_categories = Category.objects.annotate(
        annotated_event_count=Count('events')
    ).order_by('-annotated_event_count')[:5]

    # Encuestas más activas
    active_survey_list = Survey.objects.filter(status='active').annotate(
        response_count=Count('responses', filter=Q(
            responses__status='completed'))
    ).order_by('-response_count')[:5]

    context = {
        'total_events': total_events,
        'published_events': published_events,
        'upcoming_events': upcoming_events,
        'finished_events': finished_events,
        'total_registrations': total_registrations,
        'pending_registrations': pending_registrations,
        'accepted_registrations': accepted_registrations,
        'total_payments': total_payments,
        'pending_payments': pending_payments,
        'verified_payments': verified_payments,
        'total_surveys': total_surveys,
        'active_surveys': active_surveys,
        'total_survey_responses': total_survey_responses,
        'completed_survey_responses': completed_survey_responses,
        'events_with_surveys': events_with_surveys,
        'popular_events': popular_events,
        'popular_categories': popular_categories,
        'active_survey_list': active_survey_list,
    }

    return render(request, 'dashboard/events/statistics.html', context)

# ===== VISTAS DE GESTIÓN DE ENCUESTAS =====


@manager_required
def survey_list(request):
    """
    Lista de encuestas para el dashboard.
    """
    surveys = Survey.objects.select_related(
        'created_by').prefetch_related('questions').all()

    # Filtros
    status = request.GET.get('status')
    if status:
        surveys = surveys.filter(status=status)

    search = request.GET.get('search')
    if search:
        surveys = surveys.filter(
            Q(title__icontains=search) |
            Q(description__icontains=search)
        )

    # Estadísticas
    total_surveys = surveys.count()
    active_surveys = surveys.filter(status='active').count()
    draft_surveys = surveys.filter(status='draft').count()

    # Paginación
    paginator = Paginator(surveys, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'surveys': page_obj,
        'total_surveys': total_surveys,
        'active_surveys': active_surveys,
        'draft_surveys': draft_surveys,
    }

    return render(request, 'dashboard/events/survey_list.html', context)


@manager_required
def survey_create(request):
    """
    Crear nueva encuesta.
    """
    if request.method == 'POST':
        form = SurveyForm(request.POST)
        if form.is_valid():
            survey = form.save(commit=False)
            survey.created_by = request.user
            survey.save()

            messages.success(
                request, f'Encuesta "{survey.title}" creada exitosamente.')
            return redirect('dashboard:survey_detail', pk=survey.pk)
    else:
        form = SurveyForm()

    context = {
        'form': form,
        'title': 'Nueva Encuesta',
        'action': 'Crear Encuesta'
    }

    return render(request, 'dashboard/events/survey_form.html', context)


@manager_required
def survey_detail(request, pk):
    """
    Detalle de encuesta.
    """
    survey = get_object_or_404(
        Survey.objects.select_related(
            'created_by').prefetch_related('questions__options'),
        pk=pk
    )

    # Estadísticas de respuestas
    total_responses = survey.responses.count()
    completed_responses = survey.responses.filter(status='completed').count()
    opened_responses = survey.responses.filter(status='opened').count()
    expired_responses = survey.responses.filter(status='expired').count()

    # Eventos que usan esta encuesta
    events_using_survey = survey.events.all()

    context = {
        'survey': survey,
        'total_responses': total_responses,
        'completed_responses': completed_responses,
        'opened_responses': opened_responses,
        'expired_responses': expired_responses,
        'events_using_survey': events_using_survey,
    }

    return render(request, 'dashboard/events/survey_detail.html', context)


@manager_required
def survey_update(request, pk):
    """
    Editar encuesta.
    """
    survey = get_object_or_404(Survey, pk=pk)

    if request.method == 'POST':
        form = SurveyForm(request.POST, instance=survey)
        if form.is_valid():
            form.save()
            messages.success(
                request, f'Encuesta "{survey.title}" actualizada exitosamente.')
            return redirect('dashboard:survey_detail', pk=survey.pk)
    else:
        form = SurveyForm(instance=survey)

    context = {
        'form': form,
        'survey': survey,
        'title': 'Editar Encuesta',
        'action': 'Guardar Cambios'
    }

    return render(request, 'dashboard/events/survey_form.html', context)


@manager_required
def survey_delete(request, pk):
    """
    Eliminar encuesta.
    """
    survey = get_object_or_404(Survey, pk=pk)

    if request.method == 'POST':
        survey_title = survey.title
        survey.delete()
        messages.success(
            request, f'Encuesta "{survey_title}" eliminada exitosamente.')
        return redirect('dashboard:survey_list')

    context = {
        'survey': survey
    }

    return render(request, 'dashboard/events/survey_delete.html', context)


@manager_required
def survey_duplicate(request, pk):
    """
    Duplicar encuesta.
    """
    survey = get_object_or_404(Survey, pk=pk)

    if request.method == 'POST':
        new_survey = survey.duplicate(request.user)
        messages.success(
            request, f'Encuesta "{survey.title}" duplicada exitosamente.')
        return redirect('dashboard:survey_detail', pk=new_survey.pk)

    context = {
        'survey': survey
    }

    return render(request, 'dashboard/events/survey_duplicate.html', context)


@manager_required
def survey_questions(request, pk):
    """
    Gestionar preguntas de una encuesta.
    """
    survey = get_object_or_404(
        Survey.objects.prefetch_related('questions__options'), pk=pk)

    if request.method == 'POST':
        formset = SurveyQuestionFormSet(request.POST, instance=survey)
        if formset.is_valid():
            formset.save()
            messages.success(
                request, f'Preguntas de la encuesta "{survey.title}" actualizadas exitosamente.')
            return redirect('dashboard:survey_detail', pk=survey.pk)
    else:
        formset = SurveyQuestionFormSet(instance=survey)

    context = {
        'survey': survey,
        'formset': formset,
    }

    return render(request, 'dashboard/events/survey_questions.html', context)


@manager_required
def question_options(request, question_pk):
    """
    Gestionar opciones de una pregunta de opciones múltiples.
    """
    question = get_object_or_404(
        SurveyQuestion.objects.select_related('survey'), pk=question_pk)

    if question.question_type != 'multiple_choice':
        messages.error(request, 'Esta pregunta no es de opciones múltiples.')
        return redirect('dashboard:survey_questions', pk=question.survey.pk)

    if request.method == 'POST':
        formset = SurveyQuestionOptionFormSet(request.POST, instance=question)
        if formset.is_valid():
            formset.save()
            messages.success(
                request, f'Opciones de la pregunta actualizadas exitosamente.')
            return redirect('dashboard:survey_questions', pk=question.survey.pk)
        else:
            # Debug: mostrar errores del formset
            for i, form in enumerate(formset.forms):
                if form.errors:
                    messages.error(
                        request, f'Errores en formulario {i}: {form.errors}')
            if formset.non_form_errors():
                messages.error(
                    request, f'Errores del formset: {formset.non_form_errors()}')
    else:
        formset = SurveyQuestionOptionFormSet(instance=question)

    context = {
        'question': question,
        'survey': question.survey,
        'formset': formset,
    }

    return render(request, 'dashboard/events/question_options.html', context)


@manager_required
def survey_results(request, pk):
    """
    Ver resultados de una encuesta.
    """
    survey = get_object_or_404(
        Survey.objects.prefetch_related(
            'questions__options', 'responses__question_responses'),
        pk=pk
    )

    # Estadísticas generales
    total_responses = survey.responses.filter(status='completed').count()

    # Análisis por pregunta
    question_analysis = []
    for question in survey.questions.all():
        responses = question.responses.filter(
            survey_response__status='completed')

        if question.question_type == 'text':
            # Para preguntas de texto, mostrar algunas respuestas de ejemplo
            text_responses = responses.exclude(text_response='')[:5]
            analysis = {
                'question': question,
                'type': 'text',
                'response_count': responses.count(),
                'sample_responses': text_responses,
            }

        elif question.question_type == 'scale':
            # Para escalas, calcular promedio y distribución
            scale_responses = responses.exclude(scale_response__isnull=True)
            if scale_responses.exists():
                avg_rating = sum(
                    r.scale_response for r in scale_responses) / scale_responses.count()
                distribution = {}
                for i in range(1, 6):
                    distribution[i] = scale_responses.filter(
                        scale_response=i).count()
            else:
                avg_rating = 0
                distribution = {i: 0 for i in range(1, 6)}

            analysis = {
                'question': question,
                'type': 'scale',
                'response_count': responses.count(),
                'average_rating': round(avg_rating, 2),
                'distribution': distribution,
            }

        elif question.question_type == 'multiple_choice':
            # Para opciones múltiples, contar cada opción
            option_counts = {}
            for option in question.options.all():
                count = option.responses.filter(
                    survey_response__status='completed').count()
                option_counts[option.text] = count

            analysis = {
                'question': question,
                'type': 'multiple_choice',
                'response_count': responses.count(),
                'option_counts': option_counts,
            }

        question_analysis.append(analysis)

    context = {
        'survey': survey,
        'total_responses': total_responses,
        'question_analysis': question_analysis,
    }

    return render(request, 'dashboard/events/survey_results.html', context)


@manager_required
def survey_export(request, pk):
    """
    Exportar resultados de encuesta en diferentes formatos.
    """
    survey = get_object_or_404(
        Survey.objects.prefetch_related(
            'questions__options', 'responses__question_responses'),
        pk=pk
    )

    format_type = request.GET.get('format', 'csv')

    if format_type == 'csv':
        return export_survey_csv(survey)
    elif format_type == 'excel':
        return export_survey_excel(survey)
    elif format_type == 'pdf':
        return export_survey_pdf(survey)
    else:
        messages.error(request, 'Formato de exportación no válido.')
        return redirect('dashboard:survey_results', pk=pk)


def export_survey_csv(survey):
    """
    Exportar resultados de encuesta a CSV.
    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="encuesta_{survey.pk}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'

    writer = csv.writer(response)

    # Encabezados
    headers = ['Pregunta', 'Tipo', 'Respuesta', 'Participante', 'Fecha']
    writer.writerow(headers)

    # Datos
    for question in survey.questions.all():
        responses = question.responses.filter(
            survey_response__status='completed')

        for response in responses:
            if question.question_type == 'text':
                value = response.text_response
            elif question.question_type == 'scale':
                value = f"{response.scale_response}/5"
            elif question.question_type == 'multiple_choice':
                value = response.selected_option.text if response.selected_option else ''
            else:
                value = ''

            writer.writerow([
                question.text,
                question.get_question_type_display(),
                value,
                response.survey_response.registration.full_name,
                response.created_at.strftime('%d/%m/%Y %H:%M')
            ])

    return response


def export_survey_excel(survey):
    """
    Exportar resultados de encuesta a Excel.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        # Si no está instalado openpyxl, redirigir a CSV
        return export_survey_csv(survey)

    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados de Encuesta"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092",
                              end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Encabezados
    headers = ['Pregunta', 'Tipo', 'Respuesta',
               'Participante', 'Email', 'Fecha']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Datos
    row = 2
    for question in survey.questions.all():
        responses = question.responses.filter(
            survey_response__status='completed')

        for response in responses:
            if question.question_type == 'text':
                value = response.text_response
            elif question.question_type == 'scale':
                value = f"{response.scale_response}/5"
            elif question.question_type == 'multiple_choice':
                value = response.selected_option.text if response.selected_option else ''
            else:
                value = ''

            ws.cell(row=row, column=1, value=question.text)
            ws.cell(row=row, column=2, value=question.get_question_type_display())
            ws.cell(row=row, column=3, value=value)
            ws.cell(row=row, column=4,
                    value=response.survey_response.registration.full_name)
            ws.cell(row=row, column=5,
                    value=response.survey_response.registration.email)
            ws.cell(row=row, column=6,
                    value=response.created_at.strftime('%d/%m/%Y %H:%M'))
            row += 1

    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Guardar
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="encuesta_{survey.pk}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

    wb.save(response)
    return response


def export_survey_pdf(survey):
    """
    Exportar resultados de encuesta a PDF.
    """
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
    except ImportError:
        # Si no está instalado reportlab, redirigir a CSV
        return export_survey_csv(survey)

    # Crear buffer para el PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []

    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Centrado
    )
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=12,
        spaceBefore=20
    )
    normal_style = styles['Normal']

    # Título
    story.append(
        Paragraph(f"Resultados de Encuesta: {survey.title}", title_style))
    story.append(Spacer(1, 20))

    # Información general
    total_responses = survey.responses.filter(status='completed').count()
    story.append(Paragraph(
        f"<b>Total de respuestas completadas:</b> {total_responses}", normal_style))
    story.append(Paragraph(
        f"<b>Total de preguntas:</b> {survey.questions.count()}", normal_style))
    story.append(Paragraph(
        f"<b>Fecha de exportación:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}", normal_style))
    story.append(Spacer(1, 20))

    # Resultados por pregunta
    for question in survey.questions.all():
        story.append(Paragraph(f"<b>{question.text}</b>", heading_style))
        story.append(
            Paragraph(f"Tipo: {question.get_question_type_display()}", normal_style))

        responses = question.responses.filter(
            survey_response__status='completed')

        if question.question_type == 'text':
            # Mostrar algunas respuestas de texto
            text_responses = responses.exclude(text_response='')[:3]
            for response in text_responses:
                story.append(
                    Paragraph(f"• {response.text_response[:200]}...", normal_style))
                story.append(Paragraph(
                    f"  <i>Por: {response.survey_response.registration.full_name}</i>", normal_style))

        elif question.question_type == 'scale':
            # Estadísticas de escala
            scale_responses = responses.exclude(scale_response__isnull=True)
            if scale_responses.exists():
                avg_rating = sum(
                    r.scale_response for r in scale_responses) / scale_responses.count()
                story.append(
                    Paragraph(f"Promedio: {avg_rating:.2f}/5", normal_style))

                # Distribución
                distribution = {}
                for i in range(1, 6):
                    distribution[i] = scale_responses.filter(
                        scale_response=i).count()

                for rating, count in distribution.items():
                    story.append(
                        Paragraph(f"{rating} estrellas: {count} respuestas", normal_style))

        elif question.question_type == 'multiple_choice':
            # Conteo de opciones
            option_counts = {}
            for option in question.options.all():
                count = option.responses.filter(
                    survey_response__status='completed').count()
                option_counts[option.text] = count

            for option_text, count in option_counts.items():
                story.append(
                    Paragraph(f"• {option_text}: {count} respuestas", normal_style))

        story.append(Spacer(1, 15))

    # Construir PDF
    doc.build(story)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="encuesta_{survey.pk}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'

    return response


@manager_required
def send_surveys(request, event_pk):
    """
    Enviar encuestas a los participantes de un evento.
    """
    event = get_object_or_404(
        Event.objects.select_related('survey'), pk=event_pk)

    if not event.survey:
        messages.error(request, 'Este evento no tiene una encuesta asignada.')
        return redirect('dashboard:event_detail', pk=event.pk)

    if not event.send_survey:
        messages.error(
            request, 'El envío de encuestas no está habilitado para este evento.')
        return redirect('dashboard:event_detail', pk=event.pk)

    if request.method == 'POST':
        # Crear respuestas de encuesta para participantes
        created_count = create_survey_responses_for_event(event)

        # Obtener las respuestas creadas y enviar emails
        survey_responses = SurveyResponse.objects.filter(
            survey=event.survey,
            event=event,
            status='sent'
        )

        sent_count = 0
        for survey_response in survey_responses:
            if send_survey_invitation_email(survey_response):
                sent_count += 1

        messages.success(
            request, f'Se enviaron {sent_count} encuestas exitosamente.')
        return redirect('dashboard:event_detail', pk=event.pk)

    # Mostrar vista previa de envío
    registrations_to_send = event.registrations.filter(
        status='accepted',
        survey_responses__isnull=True
    )

    already_sent = event.registrations.filter(
        status='accepted',
        survey_responses__isnull=False
    )

    context = {
        'event': event,
        'registrations_to_send': registrations_to_send,
        'already_sent': already_sent,
        'total_to_send': registrations_to_send.count(),
        'total_sent': already_sent.count(),
    }

    return render(request, 'dashboard/events/send_surveys.html', context)
